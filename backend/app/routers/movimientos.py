from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import List, Optional
from datetime import date
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from app.database import get_db
from app.models.movimiento import Movimiento
from app.models.grupo import Grupo
from app.models.tipo import Tipo
from app.schemas.movimiento import MovimientoCreate, MovimientoUpdate, MovimientoOut
from app.auth.jwt import get_current_user

router = APIRouter(prefix="/api/movimientos", tags=["movimientos"])


def recalcular_saldos(db: Session):
    # Orden por ID (inserción), no por fecha — los registros pueden
    # ingresarse fuera de orden cronológico y el saldo acumulado
    # debe respetar el orden en que fueron registrados.
    movimientos = db.query(Movimiento).order_by(Movimiento.id).all()
    saldo = 0.0
    for m in movimientos:
        saldo += (m.monto_ingreso or 0) - (m.monto_egreso or 0)
        m.saldo = round(saldo, 2)
    db.commit()


@router.get("/", response_model=List[MovimientoOut])
def listar(
    skip: int = 0,
    limit: int = 100,
    grupo_id: Optional[int] = None,
    tipo_id: Optional[int] = None,
    fecha_desde: Optional[date] = None,
    fecha_hasta: Optional[date] = None,
    desc: bool = True,   # True = más recientes primero (default)
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    q = db.query(Movimiento)
    if grupo_id:
        q = q.filter(Movimiento.grupo_id == grupo_id)
    if tipo_id:
        q = q.filter(Movimiento.tipo_id == tipo_id)
    if fecha_desde:
        q = q.filter(Movimiento.fecha >= fecha_desde)
    if fecha_hasta:
        q = q.filter(Movimiento.fecha <= fecha_hasta)
    order = Movimiento.id.desc() if desc else Movimiento.id
    return q.order_by(order).offset(skip).limit(limit).all()


@router.post("/", response_model=MovimientoOut, status_code=201)
def crear(data: MovimientoCreate, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    m = Movimiento(**data.model_dump(), usuario_id=current_user.id)
    db.add(m)
    db.commit()
    recalcular_saldos(db)
    db.refresh(m)
    return m


@router.put("/{mov_id}", response_model=MovimientoOut)
def actualizar(mov_id: int, data: MovimientoUpdate, db: Session = Depends(get_db), _=Depends(get_current_user)):
    m = db.query(Movimiento).filter(Movimiento.id == mov_id).first()
    if not m:
        raise HTTPException(status_code=404, detail="Movimiento no encontrado")
    for field, val in data.model_dump(exclude_unset=True).items():
        setattr(m, field, val)
    db.commit()
    recalcular_saldos(db)
    db.refresh(m)
    return m


@router.delete("/{mov_id}")
def eliminar(mov_id: int, db: Session = Depends(get_db), _=Depends(get_current_user)):
    m = db.query(Movimiento).filter(Movimiento.id == mov_id).first()
    if not m:
        raise HTTPException(status_code=404, detail="Movimiento no encontrado")
    db.delete(m)
    db.commit()
    recalcular_saldos(db)
    return {"detail": "Eliminado"}


@router.get("/exportar/excel")
def exportar_excel(
    grupo_id: Optional[int] = None,
    tipo_id: Optional[int] = None,
    fecha_desde: Optional[date] = None,
    fecha_hasta: Optional[date] = None,
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    q = db.query(Movimiento)
    if grupo_id:
        q = q.filter(Movimiento.grupo_id == grupo_id)
    if tipo_id:
        q = q.filter(Movimiento.tipo_id == tipo_id)
    if fecha_desde:
        q = q.filter(Movimiento.fecha >= fecha_desde)
    if fecha_hasta:
        q = q.filter(Movimiento.fecha <= fecha_hasta)
    movs = q.order_by(Movimiento.id).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Movimientos"

    # Estilos
    header_fill = PatternFill("solid", fgColor="1A1A2E")
    header_font = Font(bold=True, color="E2E8F0", size=11)
    center = Alignment(horizontal="center", vertical="center")
    thin = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )
    ingreso_fill = PatternFill("solid", fgColor="F0FDF4")
    egreso_fill  = PatternFill("solid", fgColor="FFF5F5")

    headers = ["#", "Fecha", "Descripción", "Grupo", "Tipo", "Ingreso (S/)", "Egreso (S/)", "Saldo (S/)"]
    col_widths = [5, 13, 45, 22, 12, 14, 14, 14]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin
        ws.column_dimensions[cell.column_letter].width = w

    ws.row_dimensions[1].height = 22

    for i, m in enumerate(movs, 1):
        row = i + 1
        fill = ingreso_fill if m.monto_ingreso > 0 else egreso_fill
        values = [
            i,
            m.fecha.strftime("%d/%m/%Y"),
            m.descripcion,
            f"{m.grupo.icono or ''} {m.grupo.nombre}",
            m.tipo.nombre,
            m.monto_ingreso if m.monto_ingreso > 0 else "",
            m.monto_egreso if m.monto_egreso > 0 else "",
            m.saldo,
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = fill
            cell.border = thin
            cell.alignment = Alignment(vertical="center", horizontal="right" if col >= 6 else ("center" if col <= 2 else "left"))

    # Fila de totales
    total_row = len(movs) + 2
    ws.cell(row=total_row, column=3, value="TOTALES").font = Font(bold=True)
    ws.cell(row=total_row, column=6, value=sum(m.monto_ingreso for m in movs)).font = Font(bold=True, color="166534")
    ws.cell(row=total_row, column=7, value=sum(m.monto_egreso for m in movs)).font = Font(bold=True, color="991B1B")
    if movs:
        ws.cell(row=total_row, column=8, value=movs[-1].saldo).font = Font(bold=True)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = "fayni_movimientos.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/exportar/pdf")
def exportar_pdf(
    grupo_id: Optional[int] = None,
    tipo_id: Optional[int] = None,
    fecha_desde: Optional[date] = None,
    fecha_hasta: Optional[date] = None,
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    q = db.query(Movimiento)
    if grupo_id:
        q = q.filter(Movimiento.grupo_id == grupo_id)
    if tipo_id:
        q = q.filter(Movimiento.tipo_id == tipo_id)
    if fecha_desde:
        q = q.filter(Movimiento.fecha >= fecha_desde)
    if fecha_hasta:
        q = q.filter(Movimiento.fecha <= fecha_hasta)
    movs = q.order_by(Movimiento.id).all()

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.8*cm, bottomMargin=1.8*cm,
    )

    # Paleta de colores
    C_DARK   = colors.HexColor("#1A1A2E")
    C_PURPLE = colors.HexColor("#6366f1")
    C_GREEN  = colors.HexColor("#166534")
    C_RED    = colors.HexColor("#991B1B")
    C_LIGHT  = colors.HexColor("#F8FAFC")
    C_STRIPE = colors.HexColor("#F0F4FF")
    C_ING_BG = colors.HexColor("#F0FDF4")
    C_EGR_BG = colors.HexColor("#FFF5F5")
    C_BORDER = colors.HexColor("#CBD5E1")

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", fontSize=20, textColor=C_DARK, fontName="Helvetica-Bold", alignment=TA_CENTER, spaceAfter=2)
    sub_style   = ParagraphStyle("sub",   fontSize=10, textColor=C_PURPLE, fontName="Helvetica", alignment=TA_CENTER, spaceAfter=6)
    cell_style  = ParagraphStyle("cell",  fontSize=7.5, fontName="Helvetica", leading=10)
    cell_r      = ParagraphStyle("cellr", fontSize=7.5, fontName="Helvetica", leading=10, alignment=TA_RIGHT)

    from datetime import datetime
    now = datetime.now().strftime("%d/%m/%Y %H:%M")
    filtro_info = []
    if fecha_desde:
        filtro_info.append(f"Desde: {fecha_desde.strftime('%d/%m/%Y')}")
    if fecha_hasta:
        filtro_info.append(f"Hasta: {fecha_hasta.strftime('%d/%m/%Y')}")
    filtro_txt = "  |  ".join(filtro_info) if filtro_info else "Todos los movimientos"

    elements = [
        Paragraph("Fayni · Gestión de Gastos Familiares", title_style),
        Paragraph(f"Reporte de Movimientos  —  {filtro_txt}", sub_style),
        Paragraph(f"Generado: {now}", ParagraphStyle("gen", fontSize=8, textColor=colors.HexColor("#94A3B8"), alignment=TA_CENTER)),
        Spacer(1, 0.3*cm),
        HRFlowable(width="100%", thickness=1.5, color=C_PURPLE, spaceAfter=6),
    ]

    # Tabla
    col_widths_pdf = [1*cm, 2.2*cm, 7.5*cm, 3.2*cm, 2.2*cm, 2.5*cm, 2.5*cm, 2.5*cm]
    header = [
        Paragraph("<b>#</b>", ParagraphStyle("hc", fontSize=8, fontName="Helvetica-Bold", textColor=colors.white, alignment=TA_CENTER)),
        Paragraph("<b>Fecha</b>", ParagraphStyle("hc", fontSize=8, fontName="Helvetica-Bold", textColor=colors.white, alignment=TA_CENTER)),
        Paragraph("<b>Descripción</b>", ParagraphStyle("hc", fontSize=8, fontName="Helvetica-Bold", textColor=colors.white)),
        Paragraph("<b>Grupo</b>", ParagraphStyle("hc", fontSize=8, fontName="Helvetica-Bold", textColor=colors.white)),
        Paragraph("<b>Tipo</b>", ParagraphStyle("hc", fontSize=8, fontName="Helvetica-Bold", textColor=colors.white, alignment=TA_CENTER)),
        Paragraph("<b>Ingreso S/</b>", ParagraphStyle("hc", fontSize=8, fontName="Helvetica-Bold", textColor=colors.white, alignment=TA_RIGHT)),
        Paragraph("<b>Egreso S/</b>", ParagraphStyle("hc", fontSize=8, fontName="Helvetica-Bold", textColor=colors.white, alignment=TA_RIGHT)),
        Paragraph("<b>Saldo S/</b>", ParagraphStyle("hc", fontSize=8, fontName="Helvetica-Bold", textColor=colors.white, alignment=TA_RIGHT)),
    ]

    data = [header]
    row_colors = []
    for i, m in enumerate(movs, 1):
        bg = C_ING_BG if m.monto_ingreso > 0 else C_EGR_BG
        row_colors.append((i, bg))
        ingreso_txt = f"<font color='#166534'><b>{m.monto_ingreso:.2f}</b></font>" if m.monto_ingreso > 0 else ""
        egreso_txt  = f"<font color='#991B1B'><b>{m.monto_egreso:.2f}</b></font>" if m.monto_egreso > 0 else ""
        saldo_color = "#166534" if (m.saldo or 0) >= 0 else "#991B1B"
        data.append([
            Paragraph(str(i), ParagraphStyle("nc", fontSize=7.5, alignment=TA_CENTER)),
            Paragraph(m.fecha.strftime("%d/%m/%Y"), ParagraphStyle("dc", fontSize=7.5, alignment=TA_CENTER)),
            Paragraph(m.descripcion[:70], cell_style),
            Paragraph(m.grupo.nombre, cell_style),
            Paragraph(m.tipo.nombre, ParagraphStyle("tc", fontSize=7.5, alignment=TA_CENTER)),
            Paragraph(ingreso_txt, cell_r),
            Paragraph(egreso_txt, cell_r),
            Paragraph(f"<font color='{saldo_color}'><b>{(m.saldo or 0):.2f}</b></font>", cell_r),
        ])

    # Fila de totales
    total_ing = sum(m.monto_ingreso for m in movs)
    total_egr = sum(m.monto_egreso for m in movs)
    saldo_final = movs[-1].saldo if movs else 0
    saldo_color = "#166534" if saldo_final >= 0 else "#991B1B"
    data.append([
        "", "",
        Paragraph("<b>TOTALES</b>", ParagraphStyle("tot", fontSize=8, fontName="Helvetica-Bold")),
        "", "",
        Paragraph(f"<font color='#166534'><b>{total_ing:.2f}</b></font>", cell_r),
        Paragraph(f"<font color='#991B1B'><b>{total_egr:.2f}</b></font>", cell_r),
        Paragraph(f"<font color='{saldo_color}'><b>{saldo_final:.2f}</b></font>", cell_r),
    ])

    table = Table(data, colWidths=col_widths_pdf, repeatRows=1)
    ts = TableStyle([
        # Header
        ("BACKGROUND", (0, 0), (-1, 0), C_DARK),
        ("ROWBACKGROUND", (0, 0), (-1, 0), C_DARK),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("GRID", (0, 0), (-1, -1), 0.4, C_BORDER),
        # Totals row
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#EEF2FF")),
        ("LINEABOVE", (0, -1), (-1, -1), 1.2, C_PURPLE),
    ])
    # Filas con color alterno/ingreso/egreso
    for row_i, bg in row_colors:
        ts.add("BACKGROUND", (0, row_i), (-1, row_i), bg)

    table.setStyle(ts)
    elements.append(table)

    # Resumen final
    elements += [
        Spacer(1, 0.4*cm),
        HRFlowable(width="100%", thickness=0.5, color=C_BORDER),
        Spacer(1, 0.2*cm),
        Paragraph(
            f"<b>Total Ingresos:</b> <font color='#166534'>S/ {total_ing:.2f}</font>    "
            f"<b>Total Egresos:</b> <font color='#991B1B'>S/ {total_egr:.2f}</font>    "
            f"<b>Saldo Final:</b> <font color='{saldo_color}'>S/ {saldo_final:.2f}</font>    "
            f"<b>Registros:</b> {len(movs)}",
            ParagraphStyle("summary", fontSize=9, fontName="Helvetica", alignment=TA_CENTER, textColor=C_DARK),
        ),
    ]

    doc.build(elements)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=fayni_movimientos.pdf"},
    )
